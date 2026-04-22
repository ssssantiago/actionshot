import openpyxl
import rpakit


def run(planilha_path: str) -> dict:
    """Read rows from a spreadsheet and register each client."""
    app = rpakit.connect(title="Sistema Cadastro - Lote")

    # -- Load spreadsheet --
    wb = openpyxl.load_workbook(planilha_path, read_only=True)
    ws = wb["Clientes"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # -- Selectors --
    novo_sel = rpakit.Selector(automation_id="btnNovo")
    nome_sel = rpakit.Selector(automation_id="txtNome")
    cpf_sel = rpakit.Selector(automation_id="txtCPF")
    email_sel = rpakit.Selector(automation_id="txtEmail")
    salvar_sel = rpakit.Selector(automation_id="btnSalvar")
    sucesso_sel = rpakit.Selector(automation_id="lblSucesso")

    total_cadastrados = 0

    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_cells))

        # Click "Novo" to start a new record
        rpakit.wait_for(novo_sel, timeout_ms=5000)
        rpakit.click(novo_sel)

        # Wait for the form to load
        rpakit.wait_for(nome_sel, timeout_ms=5000)

        # Fill fields
        rpakit.fill(nome_sel, str(row.get("nome", "")))
        rpakit.wait_for(cpf_sel, timeout_ms=3000)
        rpakit.fill(cpf_sel, str(row.get("cpf", "")))
        rpakit.wait_for(email_sel, timeout_ms=3000)
        rpakit.fill(email_sel, str(row.get("email", "")))

        # Save
        rpakit.wait_for(salvar_sel, timeout_ms=5000)
        rpakit.click(salvar_sel)

        # Wait for confirmation
        rpakit.wait_for(sucesso_sel, timeout_ms=5000)
        total_cadastrados += 1

    wb.close()
    return {"total_cadastrados": total_cadastrados}


if __name__ == "__main__":
    result = run(planilha_path=r"C:\Dados\clientes.xlsx")
    print(f"Total cadastrados: {result['total_cadastrados']}")
